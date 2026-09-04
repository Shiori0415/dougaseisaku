# -*- coding: utf-8 -*-
"""撮影機材のアップグレード案を、最小構成とは別ファイルで書き出す。
   実行: python3 tools/build_xlsx_upgrade.py
     pdf/03_撮影機材_アップグレード案.xlsx
   カメラはお手持ちのものを使う前提。01（最小構成）との差額で比較できるようにしている。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
JP = "Meiryo"
INK, MUTED, ACCENT = "1A1A1A", "5B6266", "A8672A"
LINE, HEAD_FILL = "D6DADB", "E4E7E7"
INPUT_FILL, KEEP_FILL, UP_FILL = "FFF6D9", "F2F4F4", "E7F0EA"

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# (品目, 現行案, 現行価格, アップグレード案, 上げた価格, 何が変わるか, 優先度, 商品ページURL)
#   優先度 A＝いちばん効く／B＝画づくりが変わる／—＝据え置き
ROWS = [
    ("マイク（追加）", "─（ガンマイクのみ）", 0,
     "DJI Mic 2（送信機2＋受信機＋充電ケース）", 34339,
     "インタビュー4人分の声が別物になる。服に留めるので口元30cmで録れ、ガンマイクの「離れた音」ではなくなる。32bitフロート内部収録で音割れもしない",
     "A", "https://www.amazon.co.jp/dp/B0CFZX734J"),
    ("三脚（雲台付き）", "Velbon EX-440", 3012,
     "Manfrotto Befree live アルミニウムT三脚 ビデオ雲台キット（MVKBFRT-LIVE）", 30800,
     "フルード雲台になり、パン・ティルトがなめらかに止まる。約3,000円の三脚は「置いて固定」はできるが、動かしながら撮るとカクつく",
     "A", "https://www.ginichi.com/shop/products/detail.php?product_id=118199"),
    ("照明（LED1灯）", "Neewer 二色660 LED", 10949,
     "Godox SL-60IID ＋ ソフトボックス", 25000,
     "60Wクラスの単灯になり、③の「黒背景に一灯」で光の芯が出る。ソフトボックスで影の硬さも調整できる。Bowensマウントなので後から拡張できる",
     "B", "https://item.rakuten.co.jp/phototiroya/268-00/"),
    ("マクロレンズ", "SONY E 30mm F3.5 Macro", 27800,
     "SIGMA 70mm F2.8 DG MACRO Art（ソニーE・中古）", 52000,
     "30mmは被写体に3cmまで寄る必要があり、カメラの影が革に落ちる。70mmなら離れて同じ大きさに撮れるので、コバ塗り・刻印の寄りが破綻しない",
     "B", "https://kakaku.com/item/K0001065920/"),
    ("ジンバル（追加）", "─（三脚と手持ちで代替）", 0,
     "Zhiyun Crane M3", 20710,
     "④の街歩きと⑥の「工房を歩くカメラ」が安定する。手持ちの揺れを演出として残す方針なら、なくても成立する",
     "B", "https://www.amazon.co.jp/dp/B09J4GXV63"),
    ("音声レコーダー", "ZOOM H1essential", 12626, "同じもので十分", 12626,
     "32bitフロートで音割れしないので、これ以上のグレードは今回の用途では効果が薄い", "—", "https://www.amazon.co.jp/dp/B0CSL4PXDV"),
    ("ガンマイク", "RODE VideoMic GO II", 16364, "同じもので十分", 16364,
     "作業音（刃・ミシン・コバ磨き）を狙うのがこのマイクの役割。声はDJI Mic 2に任せるので据え置き", "—", "https://www.amazon.co.jp/dp/B09MRLGL7G"),
    ("レフ板", "EMART 丸レフ板 60cm 5in1", 1099, "同じもので十分", 1099, "上げても画に差が出ない", "—", "https://www.amazon.co.jp/dp/B0CHNY19K3"),
    ("黒スチレンボード", "A2・5mm厚 ×2枚", 1122, "同じもので十分", 1122, "上げても画に差が出ない", "—", "https://www.signmall.jp/item/10899910261.html"),
    ("アクリル板", "PULUZ 反射板 40cm（黒）", 3000, "同じもので十分", 3000, "上げても画に差が出ない", "—", "https://www.amazon.co.jp/dp/B0C49DT18X"),
    ("SDカード（任意）", "SanDisk Extreme 128GB V30", 2500, "同じもので十分", 2500,
     "4K動画ならV30で足りる。カメラのカード形式に合わせて選ぶ", "—", "https://www.amazon.co.jp/s?k=SanDisk+Extreme+SD+128GB+V30"),
]

NOTES = [
    "価格について",
    "・DJI Mic 2（送信機2＋受信機＋充電ケース）＝ Amazon実売34,339円（セール時はさらに下がります）。",
    "・Manfrotto Befree live アルミニウムT三脚 ビデオ雲台キット（MVKBFRT-LIVE）＝ 銀一 30,800円／System5 33,048円／エディオン 40,800円。表では30,800円を採用。",
    "・Godox SL-60IID ＋ ソフトボックス ＝ 約25,000円（セット構成により変動するため概算）。",
    "・SIGMA 70mm F2.8 DG MACRO Art（ソニーE用）＝ 中古 50,100〜53,700円。新品は在庫が少なく、価格.comに新品価格の登録がありません。",
    "　純正の SONY FE 90mm F2.8 Macro G OSS は約130,000円前後（海外実売 $799〜）で、今回はさすがに過剰と判断しました。",
    "",
    "注意",
    "・マクロレンズはソニーEマウント用です。お使いのカメラが別マウントの場合は、同じ焦点距離帯（70〜100mm）のマクロレンズに読み替えてください。",
    "・中古を含む価格です。中古を避ける場合はマクロレンズの金額が上がります。",
    "・カメラ本体はお手持ちのものを使う前提のため、この表には入っていません。",
    "・「いまの価格」のうち、三脚（約3,000円）とレフ板（1,099円）は今回の再確認で実売価格を確定できませんでした。購入時にご確認ください。",
    "",
    "最小構成（カメラを除く合計 約78,500円）は別ファイル「01_撮影機材と費用.xlsx」にあります。",
]


def build(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "アップグレード案"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "BROOKLYN MUSEUM ／ 向島工房　動画制作"
    ws["A1"].font = Font(name=JP, size=9, color=MUTED)
    ws["A2"] = "撮影機材　アップグレード案"
    ws["A2"].font = Font(name=JP, size=18, bold=True, color=INK)
    ws["A3"] = "カメラはお手持ちのものを使う前提。最小構成（約78,500円）から、どこにいくら足すと何が変わるかをまとめました。"
    ws["A3"].font = Font(name=JP, size=10, color=MUTED)

    ws["A5"] = "見かた"
    ws["A5"].font = Font(name=JP, size=9, bold=True, color=ACCENT)
    ws["A6"] = ("優先度A＝いちばん効く2つ（音と、カメラの支え）。B＝画づくりが変わる。「─」の行は上げても差が出ないので据え置きです。"
                "薄い黄色のセルを書き換えると3つの合計が計算し直されます。")
    ws["A6"].font = Font(name=JP, size=9, color=MUTED)

    hr = 8
    headers = ["優先度", "品目", "いまの案", "いまの価格", "上げる場合", "上げた価格", "差額", "何が変わるか", "商品ページ"]
    widths = [8, 18, 30, 12, 38, 12, 10, 56, 30]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = Font(name=JP, size=9, bold=True, color=MUTED)
        c.fill = PatternFill("solid", fgColor=HEAD_FILL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hr].height = 24

    first = hr + 1
    for i, (item, now, now_p, up, up_p, why, pri, url) in enumerate(ROWS):
        row = first + i
        for col, v in enumerate([pri, item, now, now_p, up, up_p, None, why, url], start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(col in (3, 5, 8, 9)),
                                    horizontal="center" if col == 1 else
                                    ("right" if col in (4, 6, 7) else "left"))
            c.font = Font(name=JP, size=10, color=INK)
        ws.cell(row=row, column=7, value=f"=F{row}-D{row}")
        for col in (4, 6, 7):
            ws.cell(row=row, column=col).number_format = '#,##0;(#,##0);-'
        ws.cell(row=row, column=6).fill = PatternFill("solid", fgColor=INPUT_FILL)
        fill = UP_FILL if pri in ("A", "B") else KEEP_FILL
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row=row, column=1).font = Font(name=JP, size=10, bold=(pri in ("A", "B")), color=INK)
        if url:
            lc = ws.cell(row=row, column=9)
            lc.hyperlink = url
            lc.font = Font(name=JP, size=9, color="0563C1", underline="single")
        ws.row_dimensions[row].height = 52
    last = first + len(ROWS) - 1

    # ---- 3つの合計 ----
    r = last + 2
    ws.cell(row=r, column=2, value="3つの選択肢").font = Font(name=JP, size=12, bold=True, color=INK)
    r += 1
    for i, h in enumerate(["", "選択肢", "内容", "合計", "", "", "", ""], start=1):
        if h:
            c = ws.cell(row=r, column=i, value=h)
            c.font = Font(name=JP, size=9, bold=True, color=MUTED)
            c.fill = PatternFill("solid", fgColor=HEAD_FILL)
            c.border = BORDER
    opt_first = r + 1

    # A: 最小構成（いまの価格の合計）
    ws.cell(row=opt_first, column=2, value="① 最小構成").font = Font(name=JP, size=10, bold=True, color=INK)
    ws.cell(row=opt_first, column=3, value="01のファイルと同じ。まず撮り切ることを優先する").font = Font(name=JP, size=10, color=MUTED)
    c = ws.cell(row=opt_first, column=4, value=f"=SUM(D{first}:D{last})")
    # B: 音と支えだけ上げる（優先度Aの2行だけアップグレード価格を採用）
    a_rows = [first + i for i, r_ in enumerate(ROWS) if r_[6] == "A"]
    b_rows = [first + i for i, r_ in enumerate(ROWS) if r_[6] == "B"]
    ws.cell(row=opt_first + 1, column=2, value="② 音と支えだけ上げる").font = Font(name=JP, size=10, bold=True, color=ACCENT)
    ws.cell(row=opt_first + 1, column=3,
            value="優先度Aの2つ（DJI Mic 2・フルード雲台の三脚）を足す。いちばん効く").font = Font(name=JP, size=10, color=MUTED)
    add_a = "+".join(f"G{x}" for x in a_rows)
    ws.cell(row=opt_first + 1, column=4, value=f"=SUM(D{first}:D{last})+{add_a}")
    # C: フル
    ws.cell(row=opt_first + 2, column=2, value="③ ぜんぶ上げる").font = Font(name=JP, size=10, bold=True, color=INK)
    ws.cell(row=opt_first + 2, column=3, value="A・Bすべて。照明・マクロ・ジンバルまで").font = Font(name=JP, size=10, color=MUTED)
    add_all = "+".join(f"G{x}" for x in a_rows + b_rows)
    ws.cell(row=opt_first + 2, column=4, value=f"=SUM(D{first}:D{last})+{add_all}")

    for rr in range(opt_first, opt_first + 3):
        c = ws.cell(row=rr, column=4)
        c.number_format = '#,##0'
        c.font = Font(name=JP, size=12, bold=True, color=INK)
        c.alignment = Alignment(horizontal="right")
        for col in (2, 3, 4):
            ws.cell(row=rr, column=col).border = BORDER
        ws.row_dimensions[rr].height = 26
    ws.cell(row=opt_first + 1, column=4).font = Font(name=JP, size=12, bold=True, color=ACCENT)

    r = opt_first + 5
    for line in NOTES:
        c = ws.cell(row=r, column=1, value=line)
        head = line in ("価格について", "注意")
        c.font = Font(name=JP, size=9, bold=head, color=INK if head else MUTED)
        r += 1

    ws.freeze_panes = ws.cell(row=first, column=1)
    wb.save(path)
    return path


if __name__ == "__main__":
    out = os.path.join(ROOT, "pdf")
    os.makedirs(out, exist_ok=True)
    print(build(os.path.join(out, "03_撮影機材_アップグレード案.xlsx")))
