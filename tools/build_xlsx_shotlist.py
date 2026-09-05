# -*- coding: utf-8 -*-
"""場面別ショットリスト（Aロール）を絵コンテから自動で書き出す。
   実行: python3 tools/build_xlsx_shotlist.py → pdf/02_場面別ショットリスト.xlsx

   絵コンテ（tools/deck_data.json）を唯一の情報源にしているので、絵コンテを直せばここも直る。
   並び順は動画の番号順ではなく、同じ場所・同じ機材でまとめて撮れる順（撮影の段取り順）。
   Bロール（04）とは撮影の役割で分けている。ここに載るのは、無いとその動画が成立しないカットだけ。
"""
import os, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
JP = "Meiryo"
INK, MUTED, ACCENT = "1A1A1A", "5B6266", "A8672A"
LINE, HEAD_FILL, INPUT_FILL, GRP_FILL = "D6DADB", "E4E7E7", "FFF6D9", "F3EFE7"
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CAM = "カメラ／三脚"
MAC = "カメラ／マクロレンズ／三脚／照明1灯"
ITV = "カメラ／三脚／照明1灯／音声レコーダー／ガンマイク"

# (ページ番号, シーン番号) → (場面, セット, 機材)
MAP = {
 ("01",1):("屋外・街","自宅の玄関まわり（①）",CAM+"／レフ板"),
 ("01",2):("屋外・街","自宅の玄関まわり（①）",CAM+"／レフ板"),
 ("01",3):("屋外・街","通勤路・駅・会社前（①・同じ画角を使い回す）",CAM+"／レフ板"),
 ("01",4):("屋外・街","通勤路・駅・会社前（①・同じ画角を使い回す）",CAM+"／レフ板"),
 ("01",5):("屋外・街","通勤路・駅・会社前（①・同じ画角を使い回す）",CAM+"／レフ板"),
 ("01",6):("白台・物撮り","四色を並べる（①の締め）",CAM+"／照明1灯／白台"),
 ("02",1):("工房・裁つ","作業台・真俯瞰と横（②）",MAC+"／スマホ（音の別録り）"),
 ("02",2):("工房・裁つ","作業台・真俯瞰と横（②）",MAC+"／スマホ（音の別録り）"),
 ("02",3):("工房・裁つ","作業台・真俯瞰と横（②）",MAC+"／スマホ（音の別録り）"),
 ("02",4):("工房・縫う","作業台・真俯瞰と横（②）",MAC+"／スマホ（音の別録り）"),
 ("02",5):("工房・コバ","作業台・超マクロ（②の核）",MAC+"／スマホ（音の別録り）"),
 ("02",6):("工房・コバ","作業台・超マクロ（②の核）",MAC+"／スマホ（音の別録り）"),
 ("03",1):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("03",2):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("03",3):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("03",4):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("03",5):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("03",6):("黒バック撮影","黒スチレンボード＋一灯（③専用・別日）",MAC+"／黒スチレンボード／アクリル板／白手袋"),
 ("04",1):("屋外・街","街歩き・夕方の斜光（④）",CAM+"／レフ板"),
 ("04",2):("屋外・街","街歩き・夕方の斜光（④）",CAM+"／レフ板"),
 ("04",3):("屋外・街","街歩き・夕方の斜光（④）",CAM+"／マクロレンズ／レフ板"),
 ("04",4):("屋外・街","街歩き・夕方の斜光（④）",CAM+"／マクロレンズ／レフ板"),
 ("04",5):("屋外・街","街歩き・夕方の斜光（④）",CAM+"／レフ板"),
 ("04",6):("屋外・街","白い壁の前（④の締め）",CAM+"／レフ板"),
 ("05",1):("空撮・外観","工房の上空と外観（⑤⑥共通・晴天の午前）","ドローンまたは高所／カメラ"),
 ("05",2):("複写・資料","昔の写真・カタログの複写（⑤）",CAM+"／自然光"),
 ("05",3):("インタビュー","永尾社長（⑤・画に映る／以降は声だけ流す）",ITV),
 ("05",4):("企画・設計","机まわり・サンプルと図面（⑤⑥共通）",CAM+"／照明1灯"),
 ("05",5):("企画・設計","机まわり・CADと数値（⑤⑥共通）",CAM+"／照明1灯"),
 ("05",6):("工房・全工程","一工程一カットで流す（⑤は簡潔に）",CAM),
 ("05",7):("店舗","BROOKLYN MUSEUM 店舗（⑤⑧共通）",CAM+"／ガンマイク"),
 ("05",8):("お客様・街","受け取りと使用（⑤⑧共通）",CAM),
 ("06",1):("空撮・外観","工房の上空と外観（⑤⑥共通・晴天の午前）","ドローンまたは高所／カメラ"),
 ("06",2):("工房・全体","工房の全景と通路（⑥）",CAM+"／広角"),
 ("06",3):("工房・材料","原反と金型・型紙（⑥）",CAM),
 ("06",4):("工房・裁つ","裁断機と革包丁（⑥）",CAM),
 ("06",5):("工房・下仕事","漉き・目打ち（⑥）",MAC),
 ("06",6):("工房・縫う","ミシンと手縫い（⑥）",MAC),
 ("06",7):("工房・検品","検品台（⑥・最重要）",CAM+"／照明1灯"),
 ("06",8):("梱包","梱包場（⑥）",CAM),
 ("06",9):("締めカード","撮影不要。文字を組んで作る","撮影不要"),
 ("07",1):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／マクロレンズ／照明1灯／白台／レフ板"),
 ("07",2):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／照明1灯／白台／レフ板"),
 ("07",3):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／照明1灯／白台／レフ板"),
 ("07",4):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／照明1灯／白台／レフ板"),
 ("07",5):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／照明1灯／白台／レフ板"),
 ("07",6):("白台・物撮り","白い台・正面固定（⑦専用・三品番を同じ画角で）",CAM+"／照明1灯／白台／レフ板"),
 ("08",1):("お客様・街","受け取りと使用（⑤⑧共通）",CAM),
 ("08",2):("インタビュー","購入者（⑧）",ITV),
 ("08",3):("お客様・街","使用シーン（⑧）",CAM),
 ("08",4):("お客様・開封","買った日の再現（⑧）",CAM+"／照明1灯"),
 ("08",5):("インタビュー","購入者（⑧）",ITV),
 ("08",6):("お客様・街","締め（⑧）",CAM),
}

# 撮影の段取り順（同じ場所・同じ機材でまとめて撮れる順）
ORDER = ["インタビュー","空撮・外観","工房・全体","工房・材料","工房・裁つ","工房・下仕事",
         "工房・縫う","工房・コバ","工房・全工程","工房・検品","梱包","企画・設計","複写・資料",
         "黒バック撮影","白台・物撮り","店舗","お客様・開封","お客様・街","屋外・街","締めカード"]

# 絵コンテには出てこないが、音声として必要な収録（Aロールの一部）
EXTRA = [
 ("インタビュー","佐藤さん（②に出演・声と手元）",ITV,"②",
  "この仕事を20年続けている理由を聞く。②S6で一言だけ使う（台本なし）"),
 ("インタビュー","佐藤さん（②に出演・声と手元）",ITV,"②",
  "作業音は別録りが必要なので、インタビューと手元カットは別のセッションにする"),
 ("インタビュー","くさがやさん（⑤⑥）",ITV,"⑤⑥",
  "漉き・仕上げの仕事について一言。顔出しが難しければ手元だけのバージョンも撮る"),
 ("インタビュー","永尾社長（⑤・画に映る／以降は声だけ流す）",ITV,"⑤",
  "30分収録して、経営方針の本音を拾う。台本の読み上げにしない"),
 ("インタビュー","永尾社長（⑤・画に映る／以降は声だけ流す）",ITV,"⑤",
  "創業から今までの変遷。⑤の全編に流し続けるナレーション素材になる"),
]

HEADERS = ["No", "場面", "セット（この単位でまとめて撮る）", "使う動画", "機材", "カット内容", "撮影済", "メモ"]
WIDTHS = [5, 14, 34, 12, 34, 56, 8, 22]


def plain(s):
    s = re.sub(r"<br\s*/?>", " ", s or "")
    return re.sub(r"<[^>]+>", "", s).strip()


def main():
    data = json.load(open(os.path.join(ROOT, "tools", "deck_data.json"), encoding="utf-8"))
    MARU = "①②③④⑤⑥⑦⑧"
    rows = []
    for page in data["pages"]:
        no = page["no"]
        for si, r in enumerate(page["rows"], start=1):
            scene, time, shots = r[0], r[1], r[2]
            place, setname, gear = MAP[(no, si)]
            for hi, s in enumerate(shots, start=1):
                rows.append((place, setname, f"{MARU[int(no)-1]} {scene.split()[0]}-{hi}",
                             gear, f"{plain(s[1])}　［{time}］"))
    for e in EXTRA:
        rows.append((e[0], e[1], e[3], e[2], e[4]))
    rows.sort(key=lambda x: (ORDER.index(x[0]), x[1]))

    wb = Workbook(); ws = wb.active
    ws.title = "ショットリスト"; ws.sheet_view.showGridLines = False
    ws["A1"] = "BROOKLYN MUSEUM ／ 向島工房　動画制作"
    ws["A1"].font = Font(name=JP, size=9, color=MUTED)
    ws["A2"] = "場面別ショットリスト（Aロール）"
    ws["A2"].font = Font(name=JP, size=18, bold=True, color=INK)
    ws["A3"] = "動画の番号順ではなく、同じ場所・同じ機材でまとめて撮れる順に並べています。上から順に消化してください。"
    ws["A3"].font = Font(name=JP, size=10, color=MUTED)

    r = 5
    for t, c, b in [
        ("04 Bロール参考ショット集とのちがい", INK, True),
        ("　02（この表）＝ 被写体の動作や言葉が意味を作るカット。一つ欠けるとその動画が成立しない。当日必ず消化する。", MUTED, False),
        ("　04 　　　　　＝ それ自体は何も語らないが、編集の呼吸を作る素材。無くても成立するが、無いとカットが繋がらない。", MUTED, False),
        ("　場面や工程では分けていません。撮影の役割で分けています。同じカットが両方に載ることはありません。", ACCENT, False),
        ("　この表は絵コンテから自動で作っています。絵コンテを直せばこの表も直ります。", ACCENT, False),
    ]:
        ws.cell(row=r, column=1, value=t).font = Font(name=JP, size=9, bold=b, color=c); r += 1

    r += 1
    last = 12 + len(rows) - 1
    for col, (lbl, val, col2) in enumerate([
        ("全カット数", f"=COUNTA(F12:F{last})", INK),
        ("撮影済", f"=COUNTA(G12:G{last})", ACCENT),
        ("セット数", f"=SUMPRODUCT((C12:C{last}<>\"\")/COUNTIF(C12:C{last},C12:C{last}&\"\"))", MUTED)], start=0):
        ws.cell(row=r, column=col * 2 + 1, value=lbl).font = Font(name=JP, size=9, color=MUTED)
        ws.cell(row=r, column=col * 2 + 2, value=val).font = Font(name=JP, size=11, bold=True, color=col2)

    hr = 11
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = Font(name=JP, size=9, bold=True, color=MUTED)
        c.fill = PatternFill("solid", fgColor=HEAD_FILL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hr].height = 26

    prev_place = prev_set = None
    for n, (place, setname, use, gear, body) in enumerate(rows, start=1):
        row = hr + n
        vals = [n, place if place != prev_place else "", setname if setname != prev_set else "",
                use, gear if setname != prev_set else "", body, "", ""]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name=JP, size=9, color=INK if i == 6 else MUTED)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if i == 2 and v:
                c.fill = PatternFill("solid", fgColor=GRP_FILL)
                c.font = Font(name=JP, size=9, bold=True, color=ACCENT)
            if i == 3 and v:
                c.font = Font(name=JP, size=9, bold=True, color=INK)
            if i in (7, 8):
                c.fill = PatternFill("solid", fgColor=INPUT_FILL)
        prev_place, prev_set = place, setname
        ws.row_dimensions[row].height = 32

    ws.freeze_panes = "A12"
    ws.auto_filter.ref = f"A{hr}:H{hr + len(rows)}"
    out = os.path.join(ROOT, "pdf", "02_場面別ショットリスト.xlsx")
    wb.save(out)
    print(out, f"／ {len(rows)}カット")


if __name__ == "__main__":
    main()
