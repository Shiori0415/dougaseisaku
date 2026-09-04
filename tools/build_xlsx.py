# -*- coding: utf-8 -*-
"""Googleスプレッドシートで編集できる .xlsx を2つ書き出す。
   実行: python3 tools/build_xlsx.py
     pdf/01_撮影機材と費用.xlsx      … 社長共有用。金額はSUM式なので行を足し引きしても合計が追従する
     pdf/02_場面別ショットリスト.xlsx … 工程順のショットリスト。撮影済・メモ欄に書き込んで使う
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
JP = "Meiryo"

INK = "1A1A1A"
MUTED = "5B6266"
ACCENT = "A8672A"
LINE = "D6DADB"
HEAD_FILL = "E4E7E7"
INPUT_FILL = "FFF6D9"      # 書き込んでよいセル
EXCLUDED_FILL = "F2E2DE"

thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def title_block(ws, title, subtitle, row=1):
    ws.cell(row=row, column=1, value="BROOKLYN MUSEUM ／ 向島工房　動画制作").font = Font(name=JP, size=9, color=MUTED)
    ws.cell(row=row + 1, column=1, value=title).font = Font(name=JP, size=18, bold=True, color=INK)
    ws.cell(row=row + 2, column=1, value=subtitle).font = Font(name=JP, size=10, color=MUTED)
    return row + 4


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=JP, size=9, bold=True, color=MUTED)
        c.fill = PatternFill("solid", fgColor=HEAD_FILL)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 24


# ========================= 1. 撮影機材と費用 =========================
GEAR = [
    ("カメラ本体", "SONY ZV-E10 標準ズームレンズキット", 82800,
     "8本すべての撮影本体。瞳AF・4K・外部マイク端子つき。動画向けの現行機では実売最安クラス",
     "ソニー公式・国内正規品",
     "https://www.amazon.co.jp/dp/B09BC33J91"),
    ("スマホ", "手持ちのものを使用", 0,
     "2台目のカメラとして別アングルの押さえ、作業音の別録りに使用",
     "購入不要", ""),
    ("マクロレンズ", "SONY E 30mm F3.5 Macro（SEL30M35）", 27800,
     "刃が革に入る瞬間・コバ塗り・刻印の寄り。②③⑤の核になるカットはこのレンズがないと撮れない",
     "ソニー純正・等倍マクロ",
     "https://www.amazon.co.jp/dp/B0055MFTHM"),
    ("三脚（雲台付き）", "Velbon EX-440", 3012,
     "③⑦の固定撮影と、①⑤⑥の真俯瞰。1人で撮るのでカメラを置ける三脚は必須",
     "国内カメラ用品の老舗メーカー・入門定番",
     "https://www.amazon.co.jp/dp/B0053CEPQU"),
    ("音声レコーダー", "ZOOM H1essential", 12626,
     "永尾社長・佐藤さん・くさがやさん・購入者の4人分のインタビュー音声。⑤は社長の声だけで全編が進むため最重要",
     "ZOOM公式・評価4.5／5（32件）・32bitフロートで音割れしない",
     "https://www.amazon.co.jp/s?k=ZOOM+H1essential"),
    ("ガンマイク", "RODE VideoMic GO II", 16364,
     "カメラに載せる指向性マイク。工房の環境音の中から刃の音・ミシンの音・コバを磨く音だけを拾う。電池不要",
     "RODE公式・入門ガンマイクの定番機",
     "https://www.amazon.co.jp/dp/B0D3QSNXW9"),
    ("照明（LED1灯）", "Neewer 二色660 LEDビデオライト", 10949,
     "③「黒い布一枚とライト一灯」の要。工房の蛍光灯を全部消して撮るので、この1灯が画をつくる",
     "Neewer公式・レビュー多数の定番モデル",
     "https://www.amazon.co.jp/dp/B077Z61TPN"),
    ("レフ板", "5in1レフ板／ディフューザー 60cm", 1099,
     "1灯だけだと影が硬くなるので反対側から起こす。屋外（①④）でも手元に光を回せる",
     "レビュー件数の多い定番品",
     "https://www.amazon.co.jp/dp/B0CHNY19K3"),
    ("黒スチレンボード", "ブラックスチレンボード A2・5mm厚 ×2枚", 1122,
     "③の黒背景と、余計な光を止める遮光板。布より角が立つので黒の中に製品だけを浮かせられる",
     "看板・POP資材の定番（1枚561円）",
     "https://www.signmall.jp/item/10899910261.html"),
    ("アクリル板", "PULUZ アクリル反射板 40cm（黒）", 3000,
     "製品を映り込ませる台。③の完成品カットと⑦の白台に、鏡面のリフレクションが1枚入る",
     "PULUZ公式ショップ／価格は購入時に要確認",
     "https://www.amazon.co.jp/dp/B0C49DT18X"),
    ("SDカード", "SanDisk Extreme 128GB（V30）", 2500,
     "カメラに1枚も入っていないと撮影自体ができない。4K動画なので書き込み速度はV30以上",
     "サンディスク正規品／価格は購入時に要確認",
     "https://www.amazon.co.jp/s?k=SanDisk+Extreme+SD+128GB+V30"),
]

EXCLUDED = [
    ("ジンバル", "Zhiyun Crane M3", 20710,
     "④の街歩きと⑥の工房を歩くカットの手ブレを抑える機材。三脚と手持ちでも撮影は成立するため、必要性を感じてからで良い"),
    ("予備バッテリー", "NP-FW50 互換2個＋充電器セット", 3300,
     "1日で撮り切らず日を分ければ、本体付属の1個で回せる"),
]


def build_gear(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "機材と費用"
    ws.sheet_view.showGridLines = False

    r = title_block(ws, "撮影機材と費用",
                    "動画8本を1人で撮影する前提。機材は現在なにも所有していないため、ゼロから揃える場合の金額です。")

    ws.cell(row=r, column=1, value="編集のしかた").font = Font(name=JP, size=9, bold=True, color=ACCENT)
    ws.cell(row=r + 1, column=1,
            value="薄い黄色のセル（価格）を書き換えると、合計が自動で計算し直されます。行を足す場合は合計行のすぐ上に挿入してください。"
            ).font = Font(name=JP, size=9, color=MUTED)
    r += 3

    headers = ["品目", "商品", "価格（円）", "なぜ必要か", "評価・出どころ", "商品ページ"]
    widths = [16, 36, 12, 52, 34, 30]
    header_row(ws, r, headers, widths)
    first = r + 1

    for i, (item, name, price, why, trust, url) in enumerate(GEAR):
        row = first + i
        vals = [item, name, price, why, trust, url]
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 4, 5, 6)))
            c.font = Font(name=JP, size=10, color=INK)
        pc = ws.cell(row=row, column=3)
        pc.number_format = '#,##0;(#,##0);-'
        pc.fill = PatternFill("solid", fgColor=INPUT_FILL)   # 書き換えてよいセル
        pc.alignment = Alignment(vertical="top", horizontal="right")
        if url:
            lc = ws.cell(row=row, column=6)
            lc.hyperlink = url
            lc.font = Font(name=JP, size=9, color="0563C1", underline="single")
        ws.row_dimensions[row].height = 42

    last = first + len(GEAR) - 1
    total_row = last + 1
    tc = ws.cell(row=total_row, column=1, value="合計")
    tc.font = Font(name=JP, size=11, bold=True, color=INK)
    tc.border = BORDER
    ws.cell(row=total_row, column=2).border = BORDER
    sc = ws.cell(row=total_row, column=3, value=f"=SUM(C{first}:C{last})")
    sc.font = Font(name=JP, size=12, bold=True, color=INK)
    sc.number_format = '#,##0;(#,##0);-'
    sc.alignment = Alignment(horizontal="right")
    sc.border = BORDER
    note = ws.cell(row=total_row, column=4,
                   value="アクリル板とSDカードは概算です。発注時に確定します。")
    note.font = Font(name=JP, size=9, color=MUTED)
    note.border = BORDER
    for col in (5, 6):
        ws.cell(row=total_row, column=col).border = BORDER

    r = total_row + 3
    ws.cell(row=r, column=1, value="うちカメラ本体").font = Font(name=JP, size=10, color=MUTED)
    c = ws.cell(row=r, column=3, value=f"=C{first}")
    c.number_format = '#,##0'
    c.font = Font(name=JP, size=10, color=INK)
    c.alignment = Alignment(horizontal="right")
    ws.cell(row=r + 1, column=1, value="カメラを除いた小計").font = Font(name=JP, size=10, color=MUTED)
    c = ws.cell(row=r + 1, column=3, value=f"=C{total_row}-C{first}")
    c.number_format = '#,##0'
    c.font = Font(name=JP, size=10, color=INK)
    c.alignment = Alignment(horizontal="right")
    ws.cell(row=r + 1, column=4,
            value="カメラを社内から借りられる場合はこの金額").font = Font(name=JP, size=9, color=MUTED)

    r += 4
    ws.cell(row=r, column=1, value="今回は買わないもの").font = Font(name=JP, size=12, bold=True, color=INK)
    r += 1
    header_row(ws, r, ["品目", "商品", "価格（円）", "外した理由", "", ""], widths)
    ex_first = r + 1
    for i, (item, name, price, why) in enumerate(EXCLUDED):
        row = ex_first + i
        for col, v in enumerate([item, name, price, why], start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = BORDER
            c.fill = PatternFill("solid", fgColor=EXCLUDED_FILL)
            c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 4)))
            c.font = Font(name=JP, size=10, color="6D4238")
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).alignment = Alignment(vertical="top", horizontal="right")
        ws.row_dimensions[row].height = 34
    ex_last = ex_first + len(EXCLUDED) - 1
    row = ex_last + 1
    ws.cell(row=row, column=1, value="削減額").font = Font(name=JP, size=10, bold=True, color="6D4238")
    c = ws.cell(row=row, column=3, value=f"=SUM(C{ex_first}:C{ex_last})")
    c.number_format = '#,##0'
    c.font = Font(name=JP, size=10, bold=True, color="6D4238")
    c.alignment = Alignment(horizontal="right")

    r = row + 3
    for line in [
        "評価とサクラチェックについて",
        "・カメラとレンズはソニー純正、音声はZOOMとRODEという実績あるメーカーの公式販売ページから選びました。",
        "・無名ブランドの安価な製品はレビューが操作されている可能性があるため、音声機材からは外しています",
        "　（前回案のワイヤレスピンマイク13,999円もこれに該当するため差し替えました）。",
        "・この資料の作成環境からは sakura-checker.jp に直接アクセスできず、自動でサクラ度を判定できませんでした。",
        "　メーカー公式ページであること、レビュー件数と評価を根拠にしています。",
        "・発注前に、商品ページのURLをサクラチェッカーに貼り付けてご自身でもご確認ください（数秒で判定が出ます）。",
        "",
        "将来の余地：カメラ本体をSONY ZV-E10 II（約140,000円）に替えると合計は約219,000円。暗い工房での画質に差が出ます。",
        "撮影するカットの一覧は別ファイル「02_場面別ショットリスト.xlsx」にあります。",
        "価格は2026年9月時点の実売価格です。",
    ]:
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name=JP, size=9, bold=line.endswith("について"),
                      color=INK if line.endswith("について") else MUTED)
        r += 1

    ws.freeze_panes = ws.cell(row=first, column=1)
    wb.save(path)
    return path


# ========================= 2. 場面別ショットリスト =========================
SCENES = [
    ("1 インタビュー", [
        ("永尾社長（経営方針・声のみ使用）", "⑤", "照明1灯／音声レコーダー／三脚（固定）", [
            "経営方針について本音を拾う（30分収録・台本の読み上げにしない）",
            "創業から今までの変遷、1979年から変わっていないもの（⑤の全編ナレーション素材）",
            "画には映らない前提だが、表情も別撮りしておくと今後の別動画に転用できる（任意）",
        ]),
        ("佐藤さん（この道、20年。出演）", "②", "照明1灯／音声レコーダー／三脚（固定）", [
            "この仕事を20年続けている理由、一番大切にしていること",
            "コバ塗りへの思い入れ（②S3の核カットの背景として使う一言）",
            "作業音は別録りが必要なため、インタビューと手元カットは別セッションで",
        ]),
        ("くさがやさん", "⑤", "照明1灯／音声レコーダー／三脚（固定）", [
            "革を漉く・仕上げる仕事について一言（⑤「会社の姿」パート用）",
            "顔出しの可否が未定なら、手元だけのバージョンも保険で撮っておく",
        ]),
        ("購入者インタビュー（⑧本編）", "⑧", "照明1灯／音声レコーダー／ガンマイク／三脚（固定）", [
            "日常の中で迎える。来店・使用シーンなど（⑧S1）",
            "「なぜこれを選んだか」を聞く（⑧S2）",
            "実際に使っているシーンを見せながら答えてもらう（⑧S2）",
            "購入前に迷っていた話／買ってから変わったこと（⑧S3）",
            "表情のアップ。台本にない一言を拾う（⑧S3）",
        ]),
    ]),
    ("2 企画", [
        ("新作サンプル", "⑤", "自然光／三脚", [
            "企画中の新作サンプルが並ぶ（⑤「会社の姿」パート）",
            "デザイン画・型紙と並べたカット（企画の過程が伝わる・任意）",
        ]),
    ]),
    ("3 裁断", [
        ("工房に入る（導入）", "⑤⑥", "手持ち／ガンマイク（街の音）", [
            "工房の外観と看板。ローアングル。街の音（⑥S1）",
            "中に入る。工房の全景。ハイアングル。規模が分かる画角（⑥S1）",
            "作業台に原反が広がる。真俯瞰。扱える革の大きさを示す（⑥S1・⑤S2共通）",
        ]),
        ("裁断作業", "②⑤⑥", "マクロレンズ／三脚を逆さ設置（真俯瞰用）／ガンマイク・レコーダー", [
            "超マクロ。刃が革に入る瞬間、何を見ているか分からない距離（②S1・核カット）",
            "真俯瞰。金属定規を当てて直線を一気に引く（②S2）",
            "四十五度。角を丸く回して切る（②S2）",
            "金型が置かれ、裁断機が下りる／革包丁で細部を切る（⑤S2・⑥S2共通）",
            "向島工房の職人を一人ずつ、複数カットで（特定の一人に絞らない・⑤S2）",
        ]),
    ]),
    ("4 縫製", [
        ("ミシン・手縫い", "②⑥", "マクロレンズ／三脚／ガンマイク", [
            "真俯瞰。針が革を抜ける（②S3）",
            "ミシンで縫う（横）／手縫い。二本の針が交差する（マクロ・⑥S2）",
            "横から。目だけのアップ。顔全体は入れない（②S3）",
        ]),
        ("組み立て・金具（黒バックのスタジオ）", "③", "三脚／照明1灯／黒スチレンボード／アクリル板", [
            "黒スチレンボードとライト一灯。工房の蛍光灯は全部消す（③の基本セット）",
            "手だけ。革を折り、角を作る。腕時計・指輪は外す（③S2）",
            "白手袋で金具を持ち、取り付ける。金属が触れる音（③S2）",
        ]),
        ("コバ仕上げ", "②③⑥", "マクロレンズ／照明1灯／レフ板", [
            "コバに染料を入れる（超マクロ・③S3）",
            "コバ塗り★核。塗った側と塗っていない側が同じ画面に入る。長回し・カットを割らない（②S3）",
            "目止め → コバ磨き（⑥S2）",
            "超マクロ。コバを磨く。摩擦音（②S4）",
            "完成品が黒の中に浮かぶ。上から一灯だけ。アクリル板で映り込みを作る（③S3）",
        ]),
    ]),
    ("5 検品", [
        ("検品作業", "⑤⑥", "照明1灯／三脚を逆さ設置（真俯瞰用）", [
            "職人が縫い目と断面を目で確かめる、寄り。作業音が消える（⑥S3・⑤S2共通）",
            "検品済みの製品と道具が並ぶ、真俯瞰（⑥S3）",
            "一点を手に取り、光にかざす（⑥S3）",
        ]),
    ]),
    ("6 梱包", [
        ("包む・箱詰め", "⑥", "三脚を逆さ設置（真俯瞰用）／ガンマイク（紙の音）", [
            "不織布で包み、角を折る、真俯瞰。紙の音（⑥S4）",
            "箱に納め、テープで閉じる。ここまで見せる（⑥S4）",
            "対応技法の一覧 → 問い合わせ先とURL（テロップで、⑥の締め）",
        ]),
    ]),
    ("7 店舗", [
        ("Brooklyn店舗", "⑤", "自然光／手持ち", [
            "Brooklyn店舗の外観・店内（⑤「会社の姿」パート）",
        ]),
    ]),
    ("8 お客様", [
        ("受け渡し・使用シーン", "⑤", "自然光／手持ち", [
            "購入され、店員から手渡される瞬間（⑤S4）",
            "実際に使われている手元（⑤S4）",
        ]),
        ("車内・開封シーン", "⑧", "スマホ／音声レコーダー", [
            "乗り込む。膝の上に袋、手だけ",
            "リボンをほどく、紐を引く音（無音の一拍を作る）",
            "取り出し、窓の光にかざす。断面が光る",
            "その場でカードを差し込む。もう使っている",
            "鞄を持って歩き出す後ろ姿（⑧の締め）",
        ]),
    ]),
    ("9 （参考）商品ディテール・ライフスタイル", [
        ("超マクロ・質感カット（まとめ撮り）", "①④⑤", "マクロレンズ／三脚／照明1灯", [
            "手元の小物だけに寄る（キャメル・①S2）",
            "金具や財布への超マクロ差し込み（④S2・S3）",
            "刻印が革に沈む瞬間（超マクロ・⑤S1）",
            "窪みに光が差す（⑤S1、年号テロップを重ねる想定）",
        ]),
        ("真俯瞰・四色", "①", "三脚を逆さ設置（真俯瞰用）／照明1灯／レフ板", [
            "真俯瞰。四色が並ぶ（①S4）",
        ]),
        ("屋外・自然光の手持ち（街歩き）", "①④", "カメラ手持ち／スマホ", [
            "玄関から出て、通勤路を横移動で追う（①S1〜S3、固定カメラで服の色替わりを演出）",
            "電車・エスカレーターが視界を横切る瞬間に切り替え（①S2・S3）",
            "夕方の斜光。手持ちの揺れを少しだけ残す（④の基本、一人で完結させる）",
            "電車・看板・柱をトランジションに使う（④S1）",
            "高い場所に立つ。引き・映画的（④S3）",
            "白い壁の前に立ち、英語一文を重ねる想定（④S4）",
        ]),
        ("白台・正面固定・一灯", "⑦", "三脚／照明1灯／アクリル板／ガンマイク（入れる音）", [
            "白い台の上、正面固定。カメラは一度も動かさない（⑦の基本セット）",
            "超マクロ・モノクロでファスナーの歯（⑦S1、冒頭3秒だけモノクロ設定）",
            "3品番を同じ画角・同じ小物で順番に撮る（⑦S1〜S4、台は動かさない）",
        ]),
    ]),
]


def build_shots(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ショットリスト"
    ws.sheet_view.showGridLines = False

    r = title_block(ws, "場面別ショットリスト",
                    "動画ごとではなく、実際の工程の流れで並べています。同じ動作は1回撮れば複数本に使い回せます。")

    ws.cell(row=r, column=1, value="編集のしかた").font = Font(name=JP, size=9, bold=True, color=ACCENT)
    ws.cell(row=r + 1, column=1,
            value="薄い黄色の「撮影済」「メモ」欄に書き込んで使います。例：撮影済に「✓」、メモに「10/3撮影・要再撮」。"
                  "カットを足す場合は行をそのまま挿入してください（上の「撮影済カット数」も自動で数え直します）。"
            ).font = Font(name=JP, size=9, color=MUTED)
    r += 2

    counter_row = r
    ws.cell(row=counter_row, column=1, value="全カット数").font = Font(name=JP, size=9, color=MUTED)
    ws.cell(row=counter_row, column=3, value="撮影済カット数").font = Font(name=JP, size=9, color=MUTED)
    r += 2

    headers = ["No.", "場面", "セット", "使う動画", "機材", "カット内容", "撮影済", "メモ"]
    widths = [6, 22, 30, 10, 32, 60, 9, 26]
    header_row(ws, r, headers, widths)
    first = r + 1

    row = first
    n = 0
    for scene, subs in SCENES:
        for sub_title, videos, gear, shots in subs:
            for shot in shots:
                n += 1
                vals = [n, scene, sub_title, videos, gear, shot, "", ""]
                for col, v in enumerate(vals, start=1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.border = BORDER
                    c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 3, 5, 6, 8)),
                                            horizontal="center" if col in (1, 4, 7) else "left")
                    c.font = Font(name=JP, size=10, color=INK)
                for col in (7, 8):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=INPUT_FILL)
                ws.row_dimensions[row].height = 30
                row += 1
    last = row - 1

    ws.cell(row=counter_row, column=2, value=f"=COUNTA(F{first}:F{last})").font = Font(name=JP, size=12, bold=True, color=INK)
    ws.cell(row=counter_row, column=4, value=f"=COUNTA(G{first}:G{last})").font = Font(name=JP, size=12, bold=True, color=ACCENT)

    r = last + 3
    for line in [
        "各カットの画角・秒数は絵コンテ（00_動画8本_企画と絵コンテ）を参照してください。",
        "必要な機材と金額は別ファイル「01_撮影機材と費用.xlsx」にあります。",
        "①〜⑧は動画の番号です。9番の「商品ディテール・ライフスタイル」は工房の工程とは別日程で問題ありません。",
    ]:
        ws.cell(row=r, column=1, value=line).font = Font(name=JP, size=9, color=MUTED)
        r += 1

    ws.freeze_panes = ws.cell(row=first, column=1)
    ws.auto_filter.ref = f"A{first - 1}:H{last}"
    wb.save(path)
    return path


if __name__ == "__main__":
    out = os.path.join(ROOT, "pdf")
    os.makedirs(out, exist_ok=True)
    print(build_gear(os.path.join(out, "01_撮影機材と費用.xlsx")))
    print(build_shots(os.path.join(out, "02_場面別ショットリスト.xlsx")))
